import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar
import os


FILENAME = 'attendance.xlsx'

#오늘 날짜 시트 만들기
def make_sheet_title(year,month):
    year = str(year)
    month = str(month)
    if len(month) == 1:
        month = '0' + month
    return year+'년'+month+'월'

def make_header_row(year, month):
    days_in_month = calendar.monthrange(year, month)[1]#year, month 로 해당 월의 일수 구하기
    header = ["순번", "성명"]
    for day in range(1, days_in_month + 1):
        weekday = calendar.weekday(year, month, day) # weekday(year, month, day)로 해당 날짜의 요일 구하기
        weekday_kor = ["월", "화", "수", "목", "금", "토", "일"][weekday]
        header.append(f"{day}({weekday_kor})")
    header.append("특이사항")
    return header

def get_roster_names(wb):
    if "roster" not in wb.sheetnames:
        ws = wb.create_sheet("roster")
        ws.append(["이름", "전화번호"])
        return []
    ws = wb["roster"]
    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            names.append(row[0])
    return names
def set_sum(ws, idx):
    ws.cell(row=idx, column=1, value="총합계")
    col_idx = 3
    while True:
        if ws.cell(row=2,column=col_idx).value is None:
            formula = f'=SUM(C{idx}:{get_column_letter(col_idx-2)}{idx})'
            ws.cell(row=idx, column=col_idx-1, value=formula)
            break
        col_letter = get_column_letter(col_idx)
        #3행,col이 '특이사항'인경우 row합

        formula = f'=COUNTIF({col_letter}3:{col_letter}{idx-2}, 0)'
        ws.cell(row=idx, column=col_idx, value=formula)
        col_idx += 1
def create_base_form(year, month):
    wb = openpyxl.load_workbook(FILENAME)

    sheet_title = make_sheet_title(year, month)

    if sheet_title in wb.sheetnames:
        ws = wb[sheet_title]
        print(f"⚠️ '{sheet_title}' 시트가 이미 존재합니다. 동기화 시작...")
        new_names = get_roster_names(wb)
        before_names = []
        insert_idx = 3
        for row in ws.iter_rows(min_row=3, min_col=1, max_col=2):
            if row[0].value:
                before_names.append(row[1].value)
                insert_idx+=1
            else:
                break

        print(before_names)
        print(insert_idx)
        new_names = [name for name in new_names if name not in before_names]

        if not new_names:
            print("ℹ️ 추가할 인원 없음")
            return
        #insert_idx+1번째 줄 삭제
        ws.delete_rows(insert_idx+1,1)
        
        for name in new_names:
            if insert_idx==3:
                num = 1
            else:
                num = ws.cell(row=insert_idx-1, column=1).value + 1
            ws.cell(row=insert_idx, column=1, value=num)
            ws.cell(row=insert_idx, column=2, value=name)
            insert_idx += 1
        set_sum(ws, insert_idx+1)            

        wb.save(FILENAME)
        print(f"✅ '{sheet_title}' 시트 동기화 및 저장 완료")
        return

    ws = wb.create_sheet(title=sheet_title)
    header = make_header_row(year, month)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
    title_cell = ws.cell(row=1, column=1, value=f"경로식당 {month}월 급식 수령부 (보조금)")
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="left")

    for col, text in enumerate(header, start=1):
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 8 if col < 3 else 5
        #마지막컬럼이면 9로 설정
        if col == len(header):
            ws.column_dimensions[get_column_letter(col)].width = 9

    names = get_roster_names(wb)
    for idx, name in enumerate(names, start=1):
        ws.cell(row=2 + idx, column=1, value=idx)
        ws.cell(row=2 + idx, column=2, value=name)

    data_start_row = 3
    data_end_row = data_start_row + len(names) - 1
    total_row = data_end_row + 2
    set_sum(ws, total_row)


    wb.save(FILENAME)
    print(f"✅ '{sheet_title}' 시트 생성 완료 및 저장됨")
def attendance_check():
    while True:
        wb = openpyxl.load_workbook(FILENAME)
        user_num_name_dict = {}
        ws = wb["roster"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 이름이 있는 경우
                user_num_name_dict[str(row[1])] = row[0]  # 번호:이름 매핑
        print(user_num_name_dict)
        if not user_num_name_dict:
            print("⚠️ 등록된 인원이 없습니다. roster 시트를 확인하세요.")
            return
        
        title = make_sheet_title(datetime.now().year, datetime.now().month)
        ws = wb[title]
        user_name_row_dict = {}
        for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
            if row[0]:  # 순번이 있는 경우 행 인덱스와 매핑
                if row[0] == '총합계':
                    break
                user_name_row_dict[row[1]] = int(row[0])+2# 순번은 1부터 시작하므로 +2
        today_col = datetime.now().day + 2  # 2는 순번과 성명 열을 제외한 오프셋
        user_num = input("출석할 회원의 전화번호를 입력하세요(종료 입력시 종료): ")
        if user_num == '종료':
            print("출석 체크를 종료합니다.")
            wb.save(FILENAME)
            return
        user_name = user_num_name_dict[user_num]
        row_num = user_name_row_dict.get(user_name)
        print(f"입력한 전화번호: {user_num}, 이름: {user_name}, 행 번호: {row_num}")

        ws.cell(row=row_num, column=today_col, value=0)
        wb.save(FILENAME)
    
if __name__ == "__main__":
    now = datetime.now()
    create_base_form(now.year, now.month)
    attendance_check()