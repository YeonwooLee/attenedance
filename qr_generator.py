import qrcode
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
import os

# 1. roster 시트에서 한글·영어 이름 읽어오기
wb = load_workbook('attendance.xlsx')
sheet = wb['roster']

members = []
row = 2
while True:
    kor = sheet.cell(row=row, column=1).value
    eng = sheet.cell(row=row, column=2).value
    if not kor or not eng:
        break
    members.append({'kor': kor, 'eng': eng})
    row += 1

# 2. QR 저장 폴더 준비
output_dir = 'qrcodes'
os.makedirs(output_dir, exist_ok=True)

# 3. 한글 라벨용 폰트 설정
font_path = "C:/Windows/Fonts/malgun.ttf"
font_size = 20

for member in members:
    # QR 데이터엔 영어이름만
    data = member['eng']

    # 4. QR코드 생성
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4
    )
    qr.add_data(data)       # 영어라 ASCII 모드로 안전
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert('RGB')

    # 5. 라벨 공간 확보
    qr_w, qr_h = img.size
    new_h = qr_h + font_size + 10
    new_img = Image.new('RGB', (qr_w, new_h), 'white')
    new_img.paste(img, (0, 0))

    # 6. 한글 라벨 그리기
    draw = ImageDraw.Draw(new_img)
    font = ImageFont.truetype(font_path, font_size)
    label = member['kor']
    bbox = draw.textbbox((0,0), label, font=font)
    tx = (qr_w - (bbox[2]-bbox[0])) // 2
    ty = qr_h + (10 // 2)
    draw.text((tx, ty), label, font=font, fill='black')

    # 7. 파일로 저장
    filename = f"{member['kor']}.png"
    new_img.save(os.path.join(output_dir, filename))

print("QR 생성 완료!")
