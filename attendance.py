from openpyxl import load_workbook
from datetime import datetime

# 출석부 열기
wb = load_workbook('attendance.xlsx')
# 영어→한글 매핑 생성 (roster 시트에서)
roster = wb['roster']
eng_to_kor = {}
r = 2
while True:
    kor = roster.cell(row=r, column=1).value
    eng = roster.cell(row=r, column=2).value
    if not kor or not eng:
        break
    eng_to_kor[eng.strip()] = kor.strip()
    r += 1


# 오늘 날짜 구하기
now = datetime.now()
year = str(now.year)[2:]
month = now.month
day = now.day

# 오늘 날짜에 해당하는 시트 열기
today_sheet_name = f"{year}년 {month}월"
sheet = wb[today_sheet_name]

# 오늘 날짜 컬럼 찾기
col = None
for c in range(3, sheet.max_column + 1):  # C열부터 마지막 열까지
    header = sheet.cell(row=5, column=c).value
    if header is None:
        continue
    if '(' in str(header):
        num = int(str(header).split('(')[0])  # '1(수)' -> 1
        if num == day:
            col = c
            break

if col is None:
    raise ValueError(f"오늘 날짜({day})에 해당하는 컬럼을 찾을 수 없습니다.")

# 오늘 날짜 전부 미출석(/) 처리
# C열(3번)이 1일이므로, 오늘 날짜는 2 + day 열
start_row = 6
while True:
    # 1열(순번 컬럼)이 공백이면 끝
    if sheet.cell(row=start_row, column=1).value is None:
        break
    # 오늘 날짜 열에 "/" 입력
    sheet.cell(row=start_row, column=col).value = "/"
    start_row += 1

# 이름 찾기 함수
def find_row(name):
    row = 6
    while True:
        # 이름이 비었으면 더 이상 회원 없음
        if sheet.cell(row=row, column=2).value is None:
            return None
        if sheet.cell(row=row, column=2).value == name:
            return row
        row += 1

# 출석 진행
while True:
    name = input("출석자 이름 입력 (끝내려면 엔터): ").strip()
    if name == '':
        break
    kor_name = eng_to_kor.get(name, name)       # 영어→한글, 없으면 그대로
    row = find_row(kor_name)
    if row:
        sheet.cell(row=row, column=col).value = "0"
    else:
        print(f"{name} 을 찾을 수 없습니다.")

# 파일 저장
wb.save('attendance.xlsx')
print("출석부 저장 완료!")
